import os
import uuid
import json
import requests as http_req
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
import io
from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, flash
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config API Core ───────────────────────────────────────────────────────────
CORE_API_BASE = "https://apicore.sunshinetech.com.vn"
CORE_API_KEY  = "3K4M5P7Q8RATBUCWEXFYH2J3K5N6P7R9SATCVDWEXGZH2J4M5N6Q8R9SBU"

# ── NAPAS Bank Code → Tên ngân hàng ──────────────────────────────────────────
NAPAS_BANKS = {
    "100000": "NH SUMITOMO MITSUI BANKING CORP (SMBC)",
    "100001": "Ngân hàng Citibank Việt Nam (Citibank)",
    "157979": "Ngân hàng TMCP Sài Gòn (SCB)",
    "166888": "Ngân hàng TMCP Việt Á (VAB)",
    "180906": "Ngân hàng TMCP Quốc Tế Việt Nam (VIB)",
    "191919": "Ngân hàng TMCP An Bình (ABB)",
    "422589": "Ngân hàng TNHH Một Thành Viên CIMB Việt Nam (CIMB)",
    "452999": "Ngân hàng TMCP Xuất Nhập khẩu Việt Nam (EIB)",
    "458761": "Ngân hàng HSBC Việt Nam (HSBC)",
    "546034": "Ngân hàng số CAKE by VPBank (CAKE)",
    "546035": "Ngân hàng số Ubank by VPBank (Ubank)",
    "686868": "Ngân hàng TMCP Ngoại thương Việt Nam (VCB)",
    "801011": "Ngân hàng NONGHYUP (NHB)",
    "818188": "Ngân hàng TMCP Quốc Dân (NCB)",
    "888899": "Ngân hàng TMCP Kỹ Thương Việt Nam (TCB)",
    "888999": "Ngân hàng TNHH Indovina (IVB)",
    "963399": "Ngân hàng số UMEE by KienLongBank (UMEE)",
    "970400": "Ngân hàng TMCP Sài Gòn Công Thương (SGICB)",
    "970403": "Ngân hàng TMCP Sài Gòn Thương Tín (STB)",
    "970405": "Ngân hàng Nông nghiệp và Phát triển Nông thôn Việt Nam (AGRIBANK)",
    "970406": "Ngân hàng TMCP Đông Á (DAB)",
    "970407": "Ngân hàng TMCP Kỹ Thương Việt Nam (TCB)",
    "970408": "Ngân hàng Thương mại TNHH Một Thành Viên Dầu Khí Toàn Cầu (GPB)",
    "970409": "Ngân hàng TMCP Bắc Á (BAB)",
    "970410": "Ngân hàng TNHH Một Thành Viên Standard Chartered (SCVN)",
    "970412": "Ngân hàng TMCP Đại Chúng Việt Nam (PVCB)",
    "970414": "Ngân hàng TNHH Một Thành Viên Đại Dương (OJB)",
    "970415": "Ngân hàng TMCP Công Thương Việt Nam (CTG)",
    "970416": "Ngân hàng TMCP Á Châu (ACB)",
    "970418": "Ngân hàng Đầu tư và Phát triển Việt Nam (BIDV)",
    "970419": "Ngân hàng TMCP Quốc Dân (NCB)",
    "970421": "Ngân hàng liên doanh Việt Nga (VRB)",
    "970422": "Ngân hàng TMCP Quân Đội (MB)",
    "970423": "Ngân hàng TMCP Tiên Phong (TPB)",
    "970424": "Ngân hàng TNHH Một Thành Viên Shinhan Việt Nam (SHBVN)",
    "970425": "Ngân hàng TMCP An Bình (ABB)",
    "970426": "Ngân hàng TMCP Hàng Hải Việt Nam (MSB)",
    "970427": "Ngân hàng TMCP Việt Á (VAB)",
    "970428": "Ngân hàng TMCP Nam Á (NAB)",
    "970429": "Ngân hàng TMCP Sài Gòn (SCB)",
    "970430": "Ngân hàng TMCP Thịnh vượng và Phát triển (PGB)",
    "970431": "Ngân hàng TMCP Xuất Nhập khẩu Việt Nam (EIB)",
    "970432": "Ngân hàng TMCP Việt Nam Thịnh Vượng (VPB)",
    "970433": "Ngân hàng TMCP Việt Nam Thương Tín (VIETBANK)",
    "970434": "Ngân hàng TNHH Indovina (IVB)",
    "970436": "Ngân hàng TMCP Ngoại thương Việt Nam (VCB)",
    "970437": "Ngân hàng TMCP Phát triển TP.HCM (HDB)",
    "970438": "Ngân hàng TMCP Bảo Việt (BVB)",
    "970439": "Ngân hàng TNHH Một Thành Viên Public Việt Nam (PBVN)",
    "970440": "Ngân hàng TMCP Đông Nam Á (SEAB)",
    "970441": "Ngân hàng TMCP Quốc Tế Việt Nam (VIB)",
    "970442": "Ngân hàng TNHH Một Thành Viên Hong Leong Việt Nam (HLBVN)",
    "970443": "Ngân hàng TMCP Sài Gòn - Hà Nội (SHB)",
    "970444": "TM TNHH MTV Xây Dựng Việt Nam (CBBANK)",
    "970446": "Ngân hàng Hợp Tác Xã Việt Nam (CCF)",
    "970448": "Ngân hàng TMCP Phương Đông (OCB)",
    "970449": "Ngân hàng TMCP Bưu Điện Liên Việt (LPB)",
    "970452": "NH TMCP Kiên Long (KienlongBank)",
    "970454": "Ngân hàng TMCP Bản Việt (VCCB)",
    "970455": "Ngân hàng Công nghiệp Hàn Quốc - Chi nhánh Hà Nội (IBK-HN)",
    "970456": "Ngân hàng Industrial Bank of Korea - Chi nhánh Hồ Chí Minh (IBK-HCM)",
    "970457": "Ngân hàng Woori Bank Việt Nam (WVN)",
    "970458": "Ngân hàng UOB Việt Nam (UOB)",
    "970468": "Ngân hàng TMCP Đông Nam Á (SEAB)",
    "970488": "Ngân hàng Đầu tư và Phát triển Việt Nam (BIDV)",
    "981957": "Ngân hàng TMCP Việt Nam Thịnh Vượng (VPB)",
}

import database as db
import generator as gen

app = Flask(__name__)
app.secret_key = "hd-secret-2024"
app.config["TEMPLATES_AUTO_RELOAD"] = True

UPLOAD_DIR    = os.path.join(os.path.dirname(__file__), "uploads")
TEMPLATES_DIR = os.path.join(UPLOAD_DIR, "templates")
PDFS_DIR      = os.path.join(UPLOAD_DIR, "pdfs")

db.init_db()
db.seed_banks()
db.seed_holidays()


# ── Contracts ────────────────────────────────────────────────────────────────

@app.route("/")
def contracts():
    search        = request.args.get("q", "")
    template_code = request.args.get("template", "")
    date_from     = request.args.get("date_from", "")
    date_to       = request.args.get("date_to", "")
    page          = max(1, request.args.get("page", 1, type=int))
    per_page      = request.args.get("per_page", 15, type=int)
    if per_page not in (15, 30, 50, 100):
        per_page = 15

    items, total  = db.get_contracts(search=search, template_code=template_code,
                                     page=page, per_page=per_page,
                                     date_from=date_from, date_to=date_to)
    all_templates = db.get_templates()
    total_pages   = max(1, (total + per_page - 1) // per_page)
    page          = min(page, total_pages)
    from_row      = (page - 1) * per_page + 1 if total else 0
    to_row        = min(page * per_page, total)

    contracts_list = []
    for c in items:
        c = dict(c)
        try:
            rd = json.loads(c.get("row_data") or "{}")
        except Exception:
            rd = {}
        c["ten_kh"]   = rd.get("Tên khách hàng", "")
        c["so_hd"]    = rd.get("Số Hợp đồng vay vốn", "")
        c["so_hdtc"]  = rd.get("Số HĐ Thế chấp", "")
        c["cccd"]     = rd.get("Thông tin CMND/CCCD của KH", "")
        c["lai_suat"] = rd.get("Lãi suất", "")
        c["ky_han"]   = rd.get("Kỳ hạn theo tháng", "")
        c["ngay_gd"]  = rd.get("Ngày giao dịch", "")
        contracts_list.append(c)

    return render_template("contracts.html",
                           contracts=contracts_list,
                           templates=all_templates,
                           search=search,
                           selected_template=template_code,
                           date_from=date_from,
                           date_to=date_to,
                           page=page,
                           per_page=per_page,
                           total_pages=total_pages,
                           total=total,
                           from_row=from_row,
                           to_row=to_row)


@app.route("/contracts/<int:id>")
def contract_detail(id):
    contract = db.get_contract(id)
    if not contract:
        return redirect(url_for("contracts"))
    import re as _re
    raw       = json.loads(contract["row_data"]) if contract["row_data"] else {}
    # Bỏ các cột tự sinh tên (Col32, Col33...) và cột trống giá trị
    row_data  = {k: v for k, v in raw.items()
                 if not _re.match(r'^Col\d+$', str(k)) and str(v).strip()}
    # Dùng prefix từ DB name (ổn định, không bị ảnh hưởng khi row_data bị sửa)
    name_prefix = contract["name"].rsplit(" - ", 1)[0]
    if contract["batch_id"] and name_prefix:
        batch_contracts = db.get_contracts_same_customer(contract["batch_id"], name_prefix)
    else:
        batch_contracts = [contract]
    return render_template("contract_detail.html",
                           contract=contract,
                           row_data=row_data,
                           batch_contracts=batch_contracts)


@app.route("/contracts/<int:id>/save-data", methods=["POST"])
def save_contract_data(id):
    contract = db.get_contract(id)
    if not contract:
        flash("Không tìm thấy hợp đồng", "error")
        return redirect(url_for("contracts"))

    old_data = json.loads(contract["row_data"]) if contract["row_data"] else {}
    new_data  = {k: request.form.get(k, v) for k, v in old_data.items()}
    new_json  = json.dumps(new_data, ensure_ascii=False)

    name_prefix = contract["name"].rsplit(" - ", 1)[0]
    if contract["batch_id"] and name_prefix:
        related_ids = [c["id"] for c in db.get_contracts_same_customer(contract["batch_id"], name_prefix)]
    else:
        related_ids = [id]
    db.update_contracts_row_data(related_ids, new_json)

    flash("Đã lưu thông tin", "success")
    return redirect(url_for("contract_detail", id=id))


@app.route("/contracts/<int:id>/regenerate", methods=["POST"])
def regenerate_contract(id):
    contract = db.get_contract(id)
    if not contract:
        return jsonify({"error": "Không tìm thấy"}), 404

    row_data    = json.loads(contract["row_data"]) if contract["row_data"] else {}
    name_prefix = contract["name"].rsplit(" - ", 1)[0]

    print(f"[DEBUG] contract.name  = {repr(contract['name'])}")
    print(f"[DEBUG] name_prefix    = {repr(name_prefix)}")
    print(f"[DEBUG] batch_id       = {repr(contract['batch_id'])}")

    if contract["batch_id"] and name_prefix:
        related = db.get_contracts_same_customer(contract["batch_id"], name_prefix)
    else:
        related = [contract]

    print(f"[DEBUG] related count  = {len(related)}")
    for r in related:
        print(f"[DEBUG]   -> {repr(r['name'])}")

    errors, count = [], 0
    for c in related:
        c   = dict(c)
        tpl = db.get_template(c["template_id"])
        if not tpl:
            errors.append(f"Không tìm thấy mẫu {c['template_code']}")
            continue
        try:
            os.makedirs(os.path.dirname(c["pdf_path"]), exist_ok=True)
            gen.generate_pdf(tpl["file_path"], row_data, c["pdf_path"])
            count += 1
        except Exception as e:
            errors.append(f"Mẫu {c['template_code']}: {e}")

    return jsonify({
        "success": True, "count": count, "errors": errors,
        "_debug": {"name": contract["name"], "prefix": name_prefix,
                   "batch_id": contract["batch_id"], "related": len(related)}
    })


@app.route("/contracts/bulk-delete", methods=["POST"])
def bulk_delete_contracts():
    data = request.get_json(silent=True) or {}
    ids  = [int(i) for i in data.get("ids", []) if str(i).isdigit()]
    for cid in ids:
        contract = db.get_contract(cid)
        if contract and contract["pdf_path"] and os.path.exists(contract["pdf_path"]):
            try: os.remove(contract["pdf_path"])
            except Exception: pass
        db.delete_contract(cid)
    return jsonify({"success": True, "deleted": len(ids)})


@app.route("/contracts/<int:id>/delete", methods=["POST"])
def delete_contract(id):
    contract = db.get_contract(id)
    if contract and contract["pdf_path"] and os.path.exists(contract["pdf_path"]):
        os.remove(contract["pdf_path"])
    db.delete_contract(id)
    flash("Đã xóa hợp đồng", "success")
    return redirect(url_for("contracts"))


@app.route("/contracts/<int:id>/download")
def download_contract(id):
    contract = db.get_contract(id)
    if not contract or not contract["pdf_path"] or not os.path.exists(contract["pdf_path"]):
        flash("File PDF không tồn tại", "error")
        return redirect(url_for("contract_detail", id=id))
    name = contract["name"] or f"HopDong_{id}"
    return send_file(contract["pdf_path"], as_attachment=True,
                     download_name=f"{gen.safe_filename(name)}.pdf")


# ── Generate ─────────────────────────────────────────────────────────────────

@app.route("/generate")
def generate_page():
    templates = db.get_templates()
    return render_template("generate.html", templates=templates)


def _fmt_date(d):
    """yyyy/mm/dd hoặc yyyy-mm-dd → dd/mm/yyyy"""
    if not d:
        return ""
    d = d.replace("-", "/")
    parts = d.split("/")
    if len(parts) == 3 and len(parts[0]) == 4:
        return f"{parts[2]}/{parts[1]}/{parts[0]}"
    return d


def _bank_display(bank_code):
    row = db.get_bank_by_code(bank_code)
    if row and row["bank_name"] and row["medium_name"]:
        return f"{row['bank_name']} ({row['medium_name']})"
    if row and row["bank_name"]:
        return row["bank_name"]
    return NAPAS_BANKS.get(bank_code, bank_code)


@app.route("/api/search-customer")
def search_customer():
    q = request.args.get("q", "").strip()
    if len(q) < 2:
        return jsonify(None)

    try:
        resp = http_req.get(
            f"{CORE_API_BASE}/api/v1/coreprofile/GetProviderProfileInfoByIdCardNo",
            headers={"x-api-key": CORE_API_KEY},
            params={"idCardNo": q, "idCard": q, "isInogrUser": "true"},
            timeout=10,
            verify=False
        )
        if resp.status_code != 200:
            return jsonify(None)

        raw = resp.json()
        if not raw.get("success") or not raw.get("data"):
            return jsonify(None)

        d        = raw["data"]
        bank_info = d.get("bankInfo") or {}
        accounts  = bank_info.get("accounts") or []

        result = {
            "ma_kh":            bank_info.get("cifNo", ""),
            "ho_ten":           d.get("fullName", ""),
            "so_dt":            d.get("phoneNumber", ""),
            "email":            d.get("email", ""),
            "gioi_tinh":        d.get("gender", ""),
            "ngay_sinh":        _fmt_date(d.get("dateOfBirth")),
            "quoc_tich":        "Việt Nam",
            "so_cmnd":          d.get("idCardNo", ""),
            "ngay_cap":         _fmt_date(d.get("idCardIssuedDate")),
            "ngay_het_han":     _fmt_date(d.get("idCardExpiredDate")),
            "noi_cap":          d.get("idCardIssuedBy", ""),
            "dia_chi":          d.get("fullAddress", ""),
            "ma_tk_ck":         "",
            "noi_cap_tk_ck":    "",
            "ngay_het_han_ck":  "",
            "thu_nhap_nam":     "",
            "thue_thu_nhap_nam": "",
            "nhom_kh":          str(d.get("priorityGroup", "")),
            "trang_thai_ndtcn": "",
            "tai_khoan_ngan_hang": [
                {
                    "so_tai_kh":    acc.get("accountNo", ""),
                    "so_tai_khoan": acc.get("accountNo", ""),
                    "chu_tai_khoan": acc.get("accountName", ""),
                    "ma_nh":        acc.get("napasBankCode", ""),
                    "ngan_hang":    _bank_display(acc.get("napasBankCode", "")),
                    "ma_ngan_hang": acc.get("napasBankCode", ""),
                    "chi_nhanh":    acc.get("bankBranch", ""),
                }
                for acc in accounts
            ],
        }
        return jsonify(result)
    except Exception as e:
        print(f"[CORE API ERROR] {e}")
        return jsonify({"_error": str(e)}), 502


@app.route("/api/upload-excel", methods=["POST"])
def upload_excel():
    f = request.files.get("excel")
    if not f or not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Vui lòng chọn file Excel (.xlsx)"}), 400

    temp_path = os.path.join(UPLOAD_DIR, f"tmp_{uuid.uuid4().hex}.xlsx")
    f.save(temp_path)

    try:
        wb   = load_workbook(temp_path, read_only=True, data_only=True)
        ws   = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        os.remove(temp_path)
        return jsonify({"error": f"Lỗi đọc file: {e}"}), 400

    if not rows:
        os.remove(temp_path)
        return jsonify({"error": "File Excel trống"}), 400

    headers = [str(h) if h is not None else f"Col{i}" for i, h in enumerate(rows[0])]
    preview = [
        [str(v) if v is not None else "" for v in row]
        for row in rows[1:6]
    ]
    var_names = {h: gen.to_var_name(h) for h in headers}

    return jsonify({
        "temp_path":  temp_path,
        "headers":    headers,
        "var_names":  var_names,
        "preview":    preview,
        "total_rows": len(rows) - 1,
    })


@app.route("/api/generate", methods=["POST"])
def do_generate():
    data      = request.get_json(silent=True) or {}
    temp_path = data.get("temp_path", "")

    # Basic path traversal guard
    if not temp_path.startswith(UPLOAD_DIR) or not os.path.exists(temp_path):
        return jsonify({"error": "Session hết hạn, vui lòng upload lại"}), 400

    try:
        wb   = load_workbook(temp_path, read_only=True, data_only=True)
        ws   = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        return jsonify({"error": f"Lỗi đọc file: {e}"}), 400

    headers  = [str(h) if h is not None else f"Col{i}" for i, h in enumerate(rows[0])]
    all_tpls     = list(db.get_templates())
    template_ids = data.get("template_ids")  # list of str ids from frontend
    if template_ids:
        id_set   = set(str(x) for x in template_ids)
        all_tpls = [t for t in all_tpls if str(t["id"]) in id_set]

    # Auto-detect name column — dùng "Tên khách hàng" nếu có, không thì đánh số
    NAME_COL_CANDIDATES = ["Tên khách hàng", "Ten khach hang", "TÊN KHÁCH HÀNG"]
    name_column = next((h for h in headers if h in NAME_COL_CANDIDATES), "")

    if not all_tpls:
        return jsonify({"error": "Chưa có mẫu hợp đồng nào. Vui lòng upload mẫu trước."}), 400

    batch_id  = uuid.uuid4().hex
    batch_dir = os.path.join(PDFS_DIR, batch_id)
    os.makedirs(batch_dir, exist_ok=True)

    success, errors = 0, []

    for row_idx, row in enumerate(rows[1:], start=2):
        # Skip completely empty rows
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        row_data = {headers[i]: (str(v) if v is not None else "")
                    for i, v in enumerate(row) if i < len(headers)}

        # Contract name from chosen column
        raw_name = row_data.get(name_column, "").strip() if name_column else ""
        name     = gen.safe_filename(raw_name) if raw_name else f"HopDong_{row_idx - 1:04d}"

        # Gen PDF for EVERY template
        for tpl in all_tpls:
            tpl     = dict(tpl)
            pdf_name = f"{name}_{tpl['code']}.pdf"
            pdf_path = os.path.join(batch_dir, pdf_name)
            try:
                gen.generate_pdf(tpl["file_path"], row_data, pdf_path)
                db.add_contract(
                    name          = f"{name} - {tpl['name']}",
                    template_id   = tpl["id"],
                    template_name = tpl["name"],
                    template_code = tpl["code"],
                    pdf_path      = pdf_path,
                    row_data      = json.dumps(row_data, ensure_ascii=False),
                    batch_id      = batch_id,
                )
                success += 1
            except Exception as e:
                errors.append(f"Dòng {row_idx} / Mẫu {tpl['code']}: {e}")

    db.add_batch(batch_id, os.path.basename(temp_path), success)
    try:
        os.remove(temp_path)
    except Exception:
        pass

    return jsonify({"success": True, "count": success, "errors": errors, "batch_id": batch_id})


# ── Orders (Đặt lệnh) ────────────────────────────────────────────────────────

ORDER_FIELDS = [
    "Tên khách hàng", "Giới tính", "Thông tin CMND/CCCD của KH", "Ngày cấp", "Nơi cấp",
    "Địa chỉ liên lạc", "Số điện thoại", "Ngày tháng năm Sinh",
    "Số Hợp đồng vay vốn", "Số HĐ Thế chấp", "Giá trị hợp đồng trái phiếu",
    "Số tiền bằng chữ", "Ngày giao dịch", "Ngày bằng chữ",
    "Kỳ hạn theo tháng", "Ngày đáo hạn HĐ", "Số ngày cho vay", "Lãi suất",
    "Tài khoản KH ", "Tên tài khoản KH", "Ngân hàng - chi nhánh",
    "Số lượng trái phiếu thế chấp", "Số tiền thực nhận",
    "Mã trái phiếu (theo Văn kiện trái phiếu)",
    "Mã trái phiếu (do VSDC cấp)", "TỔ CHỨC PH", "NGÀY PHÁT HÀNH ",
    "NGÀY ĐÁO HẠN ", "ĐỊA CHỈ EMAIL ", "SỐ TKCK", "NƠI MỞ TKCK", "Phí phong tỏa",
    "Người nhận hợp đồng", "SĐT người nhận", "Địa chỉ người nhận",
]


@app.route("/orders")
def orders_page():
    search    = request.args.get("q", "")
    date_from = request.args.get("date_from", "")
    date_to   = request.args.get("date_to", "")
    page      = max(1, request.args.get("page", 1, type=int))
    per_page  = request.args.get("per_page", 15, type=int)
    if per_page not in (15, 30, 50, 100): per_page = 15

    items, total = db.get_orders(search=search, page=page, per_page=per_page,
                                 date_from=date_from, date_to=date_to)
    total_pages  = max(1, (total + per_page - 1) // per_page)
    page         = min(page, total_pages)
    from_row     = (page - 1) * per_page + 1 if total else 0
    to_row       = min(page * per_page, total)

    orders_list = []
    for o in items:
        o = dict(o)
        try: d = json.loads(o.get("data") or "{}")
        except: d = {}
        o["ten_kh"]        = d.get("Tên khách hàng", "")
        o["cccd"]          = d.get("Thông tin CMND/CCCD của KH", "")
        o["so_hd"]         = d.get("Số Hợp đồng vay vốn", "")
        o["so_hdtc"]       = d.get("Số HĐ Thế chấp", "")
        o["ngay_gd"]       = d.get("Ngày giao dịch", "")
        o["ky_han"]        = d.get("Kỳ hạn theo tháng", "")
        o["nguoi_nhan"]    = d.get("Người nhận hợp đồng", "")
        o["review_status"] = o.get("review_status") or "chua_kiem_tra"
        orders_list.append(o)

    return render_template("orders.html", orders=orders_list, search=search,
                           date_from=date_from, date_to=date_to,
                           page=page, per_page=per_page, total=total,
                           total_pages=total_pages, from_row=from_row, to_row=to_row)


@app.route("/orders/check-so-hd")
def orders_check_so_hd():
    val = request.args.get("val", "").strip()
    if not val:
        return jsonify({"exists": False})
    row = db.find_order_by_so_hd(val)
    if row:
        try: d = json.loads(row["data"] or "{}")
        except: d = {}
        return jsonify({"exists": True, "order_id": row["id"],
                        "ten_kh": d.get("Tên khách hàng", "")})
    return jsonify({"exists": False})


@app.route("/orders/export")
def orders_export():
    search    = request.args.get("q", "")
    date_from = request.args.get("date_from", "")
    date_to   = request.args.get("date_to", "")
    all_items, _ = db.get_orders(search=search, page=1, per_page=99999,
                                 date_from=date_from, date_to=date_to)

    rows = []
    for o in all_items:
        o = dict(o)
        try: d = json.loads(o.get("data") or "{}")
        except: d = {}
        rows.append({
            "ten_kh":       d.get("Tên khách hàng", ""),
            "gioi_tinh":    d.get("Giới tính", ""),
            "cccd":         d.get("Thông tin CMND/CCCD của KH", ""),
            "so_hd":        d.get("Số Hợp đồng vay vốn", ""),
            "so_hdtc":      d.get("Số HĐ Thế chấp", ""),
            "ngay_gd":      d.get("Ngày giao dịch", ""),
            "ky_han":       d.get("Kỳ hạn theo tháng", ""),
            "so_tien":      d.get("Giá trị hợp đồng trái phiếu", ""),
            "nguoi_nhan":   d.get("Người nhận hợp đồng", ""),
            "sdt_nhan":     d.get("SĐT người nhận", ""),
            "dc_nhan":      d.get("Địa chỉ người nhận", ""),
            "review":       "Đã KT" if o.get("review_status") == "da_kiem_tra" else "Chưa KT",
            "status":       "Đã gen HĐ" if o.get("status") == "generated" else "Nháp",
            "created_at":   (o.get("created_at") or "")[:16].replace("T", " "),
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách lệnh"

    HEADERS = [
        "STT", "Tên khách hàng", "Giới tính", "CCCD/CMND",
        "Số HĐ vay vốn", "Số HĐ Thế chấp", "Ngày giao dịch", "Kỳ hạn (tháng)",
        "Giá trị HĐ", "Người nhận HĐ", "SĐT người nhận", "Địa chỉ người nhận",
        "Tình trạng KT", "Trạng thái", "Ngày tạo",
    ]
    COL_WIDTHS = [6, 28, 10, 18, 26, 22, 14, 12, 18, 22, 16, 30, 12, 14, 18]

    # title row
    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
    title_cell = ws["A1"]
    title_cell.value = "DANH SÁCH ĐẶT LỆNH"
    title_cell.font      = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # header row
    hdr_fill   = PatternFill("solid", fgColor="1E3A5F")
    hdr_font   = Font(bold=True, color="FFFFFF", size=10)
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side  = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS[ci - 1]
    ws.row_dimensions[2].height = 22

    # data rows
    even_fill = PatternFill("solid", fgColor="F5F8FF")
    data_align = Alignment(vertical="center", wrap_text=False)

    for ri, r in enumerate(rows, 1):
        row_data = [
            ri, r["ten_kh"], r["gioi_tinh"], r["cccd"],
            r["so_hd"], r["so_hdtc"], r["ngay_gd"], r["ky_han"],
            r["so_tien"], r["nguoi_nhan"], r["sdt_nhan"], r["dc_nhan"],
            r["review"], r["status"], r["created_at"],
        ]
        fill = even_fill if ri % 2 == 0 else None
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri + 2, column=ci, value=val)
            cell.alignment = data_align
            cell.border    = thin_border
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from datetime import datetime
    fname = f"danh_sach_lenh_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/orders/new", methods=["GET", "POST"])
def order_new():
    if request.method == "POST":
        data  = {f: request.form.get(f, "") for f in ORDER_FIELDS}
        seq   = db.get_next_order_seq()
        seq_s = str(seq).zfill(5)
        ma_vk = data.get("Mã trái phiếu (theo Văn kiện trái phiếu)", "")
        data["Số Hợp đồng vay vốn"] = f"{seq_s}/HDVV-KGALAXY.{ma_vk}-KSG01"
        data["Số HĐ Thế chấp"]     = f"{seq_s}/HDTC-KGALAXY.{ma_vk}-KSG01"
        oid = db.add_order(json.dumps(data, ensure_ascii=False))
        db.bump_order_seq()
        flash("Đã thêm lệnh mới", "success")
        return redirect(url_for("order_detail", id=oid))
    return render_template("order_form.html", order=None, fields=ORDER_FIELDS, data={})


@app.route("/orders/<int:id>/review", methods=["POST"])
def order_review(id):
    order = db.get_order(id)
    if not order:
        return jsonify({"error": "Không tìm thấy"}), 404
    current = order["review_status"] if "review_status" in order.keys() else "chua_kiem_tra"
    new_status = "da_kiem_tra" if current != "da_kiem_tra" else "chua_kiem_tra"
    db.update_order_review(id, new_status)
    return jsonify({"ok": True, "review_status": new_status})


@app.route("/orders/<int:id>")
def order_detail(id):
    order = db.get_order(id)
    if not order: return redirect(url_for("orders_page"))
    data = json.loads(order["data"]) if order["data"] else {}
    contracts = [dict(c) for c in db.get_contracts_by_order(id)]
    return render_template("order_detail.html", order=order, data=data,
                           fields=ORDER_FIELDS, templates=db.get_templates(),
                           contracts=contracts)


@app.route("/orders/<int:id>/edit", methods=["GET", "POST"])
def order_edit(id):
    order = db.get_order(id)
    if not order: return redirect(url_for("orders_page"))
    data = json.loads(order["data"]) if order["data"] else {}
    if request.method == "POST":
        new_data = {f: request.form.get(f, "") for f in ORDER_FIELDS}
        db.update_order(id, json.dumps(new_data, ensure_ascii=False))
        flash("Đã lưu thay đổi", "success")
        return redirect(url_for("order_detail", id=id))
    return render_template("order_form.html", order=order, fields=ORDER_FIELDS, data=data)


@app.route("/orders/<int:id>/generate", methods=["POST"])
def order_generate(id):
    order = db.get_order(id)
    if not order: return jsonify({"error": "Không tìm thấy"}), 404

    row_data     = json.loads(order["data"]) if order["data"] else {}
    template_ids = request.get_json(silent=True, force=True).get("template_ids", [])
    all_tpls     = list(db.get_templates())
    if template_ids:
        id_set   = set(str(x) for x in template_ids)
        all_tpls = [t for t in all_tpls if str(t["id"]) in id_set]
    if not all_tpls:
        return jsonify({"error": "Không có mẫu nào được chọn"}), 400

    raw_name  = row_data.get("Tên khách hàng", "").strip()
    name      = gen.safe_filename(raw_name) if raw_name else f"LeNH_{id:04d}"
    batch_id  = uuid.uuid4().hex
    batch_dir = os.path.join(PDFS_DIR, batch_id)
    os.makedirs(batch_dir, exist_ok=True)

    errors, count = [], 0
    for tpl in all_tpls:
        tpl      = dict(tpl)
        pdf_path = os.path.join(batch_dir, f"{name}_{tpl['code']}.pdf")
        try:
            gen.generate_pdf(tpl["file_path"], row_data, pdf_path)
            db.add_contract(
                name=f"{name} - {tpl['name']}", template_id=tpl["id"],
                template_name=tpl["name"], template_code=tpl["code"],
                pdf_path=pdf_path,
                row_data=json.dumps(row_data, ensure_ascii=False),
                batch_id=batch_id,
                order_id=id,
            )
            count += 1
        except Exception as e:
            errors.append(f"Mẫu {tpl['code']}: {e}")

    db.add_batch(batch_id, f"order_{id}", count)
    db.update_order(id, order["data"], status="generated")
    return jsonify({"success": True, "count": count, "errors": errors, "batch_id": batch_id})


@app.route("/orders/<int:id>/delete", methods=["POST"])
def order_delete(id):
    db.delete_order(id)
    flash("Đã xóa lệnh", "success")
    return redirect(url_for("orders_page"))


@app.route("/orders/bulk-delete", methods=["POST"])
def bulk_delete_orders():
    data = request.get_json(silent=True) or {}
    ids  = [int(i) for i in data.get("ids", []) if str(i).isdigit()]
    db.bulk_delete_orders(ids)
    return jsonify({"success": True, "deleted": len(ids)})


# ── Settings / Policies (Chính sách sản phẩm) ───────────────────────────────

@app.route("/settings")
def settings_page():
    policies = [dict(r) for r in db.get_policies()]
    return render_template("settings.html", policies=policies)


@app.route("/settings/holidays")
def holidays_page():
    search   = request.args.get("q", "")
    holidays = [dict(r) for r in db.get_holidays(search=search)]
    return render_template("holidays.html", holidays=holidays, search=search)


@app.route("/settings/holidays/new", methods=["POST"])
def holiday_new():
    data = request.get_json(silent=True) or {}
    ngay = data.get("ngay", "").strip()
    loai = data.get("loai", "").strip()
    if not ngay or not loai:
        return jsonify({"error": "Vui lòng điền đầy đủ thông tin"}), 400
    hid = db.add_holiday(ngay, loai)
    h   = dict(db.get_holiday(hid))
    return jsonify({"success": True, "holiday": h})


@app.route("/settings/holidays/<int:id>/edit", methods=["POST"])
def holiday_edit(id):
    data = request.get_json(silent=True) or {}
    ngay = data.get("ngay", "").strip()
    loai = data.get("loai", "").strip()
    if not ngay or not loai:
        return jsonify({"error": "Vui lòng điền đầy đủ thông tin"}), 400
    db.update_holiday(id, ngay, loai)
    return jsonify({"success": True})


@app.route("/settings/holidays/<int:id>/delete", methods=["POST"])
def holiday_delete(id):
    db.delete_holiday(id)
    return jsonify({"success": True})


@app.route("/settings/banks")
def banks_page():
    search = request.args.get("q", "")
    banks  = [dict(r) for r in db.get_banks(search=search)]
    return render_template("banks.html", banks=banks, search=search)


@app.route("/settings/policies", methods=["POST"])
def policy_create():
    data = request.get_json(silent=True) or {}
    ten  = data.get("ten_chinh_sach", "").strip()
    pct  = data.get("pct_the_chap", 0)
    if not ten:
        return jsonify({"error": "Tên chính sách không được để trống"}), 400
    try:
        pct = float(pct)
    except (ValueError, TypeError):
        pct = 0.0
    pid = db.add_policy(ten, pct)
    policy = dict(db.get_policy(pid))
    return jsonify({"success": True, "policy": policy})


@app.route("/settings/policies/<int:id>")
def policy_detail(id):
    policy = db.get_policy(id)
    if not policy:
        return redirect(url_for("settings_page"))
    tenors = [dict(t) for t in db.get_policy_tenors(id)]
    return render_template("policy_detail.html", policy=dict(policy), tenors=tenors)


@app.route("/settings/policies/<int:id>/edit", methods=["POST"])
def policy_edit(id):
    data = request.get_json(silent=True) or {}
    ten  = data.get("ten_chinh_sach", "").strip()
    pct  = data.get("pct_the_chap", 0)
    try:
        pct = float(pct)
    except (ValueError, TypeError):
        pct = 0.0
    db.update_policy(id, ten, pct)
    return jsonify({"success": True})


@app.route("/settings/policies/<int:id>/activate", methods=["POST"])
def policy_activate(id):
    db.update_policy_status(id, "active")
    return jsonify({"success": True})


@app.route("/settings/policies/<int:id>/close", methods=["POST"])
def policy_close(id):
    db.update_policy_status(id, "closed")
    return jsonify({"success": True})


@app.route("/settings/policies/<int:id>/delete", methods=["POST"])
def policy_delete(id):
    db.delete_policy(id)
    flash("Đã xóa chính sách", "success")
    return redirect(url_for("settings_page"))


@app.route("/settings/policies/<int:id>/tenors", methods=["POST"])
def policy_tenor_add(id):
    data    = request.get_json(silent=True) or {}
    ky_han  = data.get("ky_han", 0)
    loi_tuc = data.get("loi_tuc", 0)
    try:
        ky_han  = int(ky_han)
        loi_tuc = float(loi_tuc)
    except (ValueError, TypeError):
        return jsonify({"error": "Giá trị không hợp lệ"}), 400
    if ky_han <= 0:
        return jsonify({"error": "Kỳ hạn phải lớn hơn 0"}), 400
    tid   = db.add_policy_tenor(id, ky_han, loi_tuc)
    tenor = next((dict(t) for t in db.get_policy_tenors(id) if t["id"] == tid), {})
    return jsonify({"success": True, "tenor": tenor})


@app.route("/settings/policies/<int:id>/tenors/<int:tid>/edit", methods=["POST"])
def policy_tenor_edit(id, tid):
    data    = request.get_json(silent=True) or {}
    ky_han  = data.get("ky_han", 0)
    loi_tuc = data.get("loi_tuc", 0)
    try:
        ky_han  = int(ky_han)
        loi_tuc = float(loi_tuc)
    except (ValueError, TypeError):
        return jsonify({"error": "Giá trị không hợp lệ"}), 400
    db.update_policy_tenor(tid, ky_han, loi_tuc)
    return jsonify({"success": True})


@app.route("/settings/policies/<int:id>/tenors/<int:tid>/delete", methods=["POST"])
def policy_tenor_delete(id, tid):
    db.delete_policy_tenor(tid)
    return jsonify({"success": True})


# ── Bond Lots (Lô Trái Phiếu) ───────────────────────────────────────────────

@app.route("/bond-lots")
def bond_lots_page():
    search   = request.args.get("q", "")
    page     = max(1, request.args.get("page", 1, type=int))
    per_page = request.args.get("per_page", 20, type=int)
    if per_page not in (15, 20, 30, 50): per_page = 20
    items, total = db.get_bond_lots(search=search, page=page, per_page=per_page)
    total_pages  = max(1, (total + per_page - 1) // per_page)
    page         = min(page, total_pages)
    lots         = [dict(r) for r in items]
    return render_template("bond_lots.html", lots=lots, search=search,
                           page=page, per_page=per_page, total=total,
                           total_pages=total_pages,
                           from_row=(page-1)*per_page+1 if total else 0,
                           to_row=min(page*per_page, total))


@app.route("/bond-lots/new", methods=["GET", "POST"])
def bond_lot_new():
    if request.method == "POST":
        don_gia_raw = request.form.get("don_gia", "0").strip().replace(",", "").replace(".", "")
        db.add_bond_lot(
            ma_vk      = request.form.get("ma_vk", "").strip(),
            ma_vsdc    = request.form.get("ma_vsdc", "").strip(),
            to_chuc_ph = request.form.get("to_chuc_ph", "").strip(),
            ngay_ph    = request.form.get("ngay_ph", "").strip(),
            ngay_dh    = request.form.get("ngay_dh", "").strip(),
            don_gia    = float(don_gia_raw) if don_gia_raw else 0,
        )
        flash("Đã thêm lô trái phiếu", "success")
        return redirect(url_for("bond_lots_page"))
    return render_template("bond_lot_form.html", lot=None)


@app.route("/bond-lots/<int:id>/edit", methods=["GET", "POST"])
def bond_lot_edit(id):
    lot = db.get_bond_lot(id)
    if not lot:
        return redirect(url_for("bond_lots_page"))
    if request.method == "POST":
        don_gia_raw = request.form.get("don_gia", "0").strip().replace(",", "").replace(".", "")
        db.update_bond_lot(
            id,
            ma_vk      = request.form.get("ma_vk", "").strip(),
            ma_vsdc    = request.form.get("ma_vsdc", "").strip(),
            to_chuc_ph = request.form.get("to_chuc_ph", "").strip(),
            ngay_ph    = request.form.get("ngay_ph", "").strip(),
            ngay_dh    = request.form.get("ngay_dh", "").strip(),
            don_gia    = float(don_gia_raw) if don_gia_raw else 0,
        )
        flash("Đã lưu thay đổi", "success")
        return redirect(url_for("bond_lots_page"))
    return render_template("bond_lot_form.html", lot=lot)


@app.route("/bond-lots/<int:id>/delete", methods=["POST"])
def bond_lot_delete(id):
    db.delete_bond_lot(id)
    flash("Đã xóa lô trái phiếu", "success")
    return redirect(url_for("bond_lots_page"))


@app.route("/api/bond-lots")
def api_bond_lots():
    items, _ = db.get_bond_lots(page=1, per_page=500)
    return jsonify([dict(r) for r in items])


@app.route("/api/active-policies")
def api_active_policies():
    policies = [dict(r) for r in db.get_policies()]
    return jsonify([p for p in policies if p["status"] == "active"])


@app.route("/api/policies/<int:id>/tenors")
def api_policy_tenors_get(id):
    return jsonify([dict(r) for r in db.get_policy_tenors(id)])


@app.route("/api/holidays")
def api_holidays():
    return jsonify([r["ngay"] for r in db.get_holidays()])


@app.route("/api/next-order-seq")
def api_next_order_seq():
    return jsonify({"seq": db.get_next_order_seq()})


# ── Templates (Word mẫu) ──────────────────────────────────────────────────────

@app.route("/templates")
def templates_page():
    items = db.get_templates()
    return render_template("templates.html", templates=items)


@app.route("/templates/upload", methods=["POST"])
def upload_template():
    name = request.form.get("name", "").strip()
    code = request.form.get("code", "").strip().upper()
    f    = request.files.get("file")

    if not name or not code or not f:
        flash("Vui lòng điền đầy đủ thông tin", "error")
        return redirect(url_for("templates_page"))

    if not f.filename.lower().endswith(".docx"):
        flash("Chỉ nhận file .docx", "error")
        return redirect(url_for("templates_page"))

    if db.get_template_by_code(code):
        flash(f'Mã "{code}" đã tồn tại', "error")
        return redirect(url_for("templates_page"))

    filename  = f"{code}_{uuid.uuid4().hex[:8]}.docx"
    file_path = os.path.join(TEMPLATES_DIR, filename)
    f.save(file_path)

    db.add_template(name=name, code=code, filename=f.filename, file_path=file_path)
    flash(f'Đã thêm mẫu "{name}" (mã: {code})', "success")
    return redirect(url_for("templates_page"))


@app.route("/templates/<int:id>/download")
def download_template(id):
    tpl = db.get_template(id)
    if not tpl or not os.path.exists(tpl["file_path"]):
        flash("File mẫu không tồn tại", "error")
        return redirect(url_for("templates_page"))
    return send_file(tpl["file_path"], as_attachment=True,
                     download_name=f"{tpl['code']}_{tpl['filename']}")


@app.route("/templates/<int:id>/delete", methods=["POST"])
def delete_template(id):
    tpl = db.get_template(id)
    if tpl and os.path.exists(tpl["file_path"]):
        os.remove(tpl["file_path"])
    db.delete_template(id)
    flash("Đã xóa mẫu hợp đồng", "success")
    return redirect(url_for("templates_page"))


# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import webbrowser
    webbrowser.open("http://localhost:5000")
    app.run(host="0.0.0.0", port=5000, debug=False)
